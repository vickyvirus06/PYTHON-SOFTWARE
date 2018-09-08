import openpyxl
from openpyxl import *
from tkinter import *
import tkinter as tk
import time



# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('/home/vickyvirus/VICKYVIRUS PROGRAMS/shivdatt2018.xlsx')

# create the sheet object
sheet = wb.active


def excel():
    
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 20
    
   
    

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "SR"
    sheet.cell(row=1, column=2).value = "NAME"
    sheet.cell(row=1, column=3).value = "AMOUNT"
    sheet.cell(row=1, column=4).value = "PAID/UNPAID"


# Function to set focus (cursor)
#def focus1(event):
    # set focus on the name box
    #sr.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the amount box
    amount.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the paid_unpaid box
    paid_unpaid.focus_set()







# Function for clearing the
# contents of text entry boxes

def clear():
    
    # clear the content of text entry box
    
   # sr.delete(0, END)
    name.delete(0, END)
    amount.delete(0, END)
    
    


# Function to take data from GUI 
# window and write to an excel file
def insert():
    
    # if user not fill any entry
    # then print "empty input"
    if (#sr.get() == "" and
        name.get() == "" and
        amount.get() == "" and
        paid_unpaid.get() == ""):
        
            
        print("empty input")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        for i in range(2,sheet.max_row+2,1):
            cell=sheet.cell(row=i,column=1)
            ik=i-1
            st = str(ik)
            sheet.cell(row=i,column=1).value = st
            srnumber=int(st)
            srnoadd=srnumber+1
            srstring = str(srnoadd)
            
            
            sr.config(text=srstring)
        
        #sheet.cell(row=current_row + 1, column=1).value = sr.get()
        sheet.cell(row=current_row + 1, column=2).value = name.get()
        sheet.cell(row=current_row + 1, column=3).value = amount.get()
        
        """ sheet.cell(row=x*x,column=i+1) = amount.get()
        """
        sheet.cell(row=current_row + 1, column=4).value = paid_unpaid.get()
                # save the file
        wb.save('/home/vickyvirus/VICKYVIRUS PROGRAMS/shivdatt2018.xlsx')

        # set focus on the sr box
        name.focus_set()

        # call the clear() function
        clear()



# create a GUI window
root = Tk()

    # set the background colour of GUI window

    # set the title of GUI window
root.title("SHIVDATT MITRA MANDAL")
root.configure(bg="powder blue")


    # set the configuration of GUI window
root.geometry("1280x1024+0+0")

excel()

    # create a Form label

Tops = Frame(root,width=1280,height=500,bg="powder blue",relief=SUNKEN)
Tops.pack(side=TOP)
f1 = Frame(root,width=800,height=700,bg="powder blue",relief=SUNKEN)
f1.pack(side=TOP)
f2 = Frame(root,width=800,height=100,bg="powder blue",relief=SUNKEN)
f2.pack(side=RIGHT)

#===============Time=======================
localtime=time.asctime(time.localtime(time.time()))
#=========================Info=================================
h = Label(Tops,font=('arial',50,'bold'), text="SHIVDATT MITRA MANDAL", fg="steel blue",bg = "powder blue",bd=5,anchor='w').grid(row = 0,column = 0)
lbinfo = Label(Tops,font=('arial',20,'bold'),text = localtime,fg = "steel blue",bd = 10,anchor = 'w').grid(row = 1,column = 0)
#========================================entry=================

    
def total():
    pd=0
    unpd=0
    
    
    for i in range(2,sheet.max_row+1,1):
    
        cell=sheet.cell(row=i,column=4)
        if cell.value == "PAID":
            cell1=sheet.cell(row=i,column=3)
            pd = int(cell1.value)+pd
        else:
            cell2=sheet.cell(row=i,column=3)
            unpd = int(cell2.value)+unpd
    tl=pd+unpd
    total=str(tl)
    paid = str(pd)
    unpaid = str(unpd)
    totalamount.config(text=  '  TOTAL          '+  total)   
    paidamount.config(text=   '  PAID             '+ paid)
    unpaidamount.config(text= 'UNPAID          '+  unpaid)


def change():
    
    s_new = Label(f1,font=('arial',18,'bold'),text="SR",bd=16,fg="steel blue",bg="powder blue")
    s_new.grid(row=8,column=0)
   
    sr_new = Entry(f1,font=('arial',16,'bold'),bd=10,bg="powder blue")
    sr_new.grid(row=8,column=1)
    sr_new.focus_set()
    """sr= for i in range(sheet.max_row = i,column = 2)
    config(text = 'sheet.cell.value(column = 3 )')
    """    
    
    
    def edit():
        
        sr_new1 = sr_new.get()
        for i in range(2,sheet.max_row+1,1):
            cell_sheet=sheet.cell(row=i,column=1)
            if cell_sheet.value == sr_new1:
                cell_name = sheet.cell(row=i,column=2).value
                nameshow = Label(f1,font=('arial',20,'bold'),text="NAME", bd=10,fg="steel blue",bg="powder blue")
                nameshow.grid(row=10,column=0)
                nameshow = Label(f1,font=('arial',20,'bold'),text=cell_name, bd=10,fg="purple",bg="powder blue")
                nameshow.grid(row=10,column=1)
                pd_ud = StringVar()
                pn = Label(f1,font=('arial',18,'bold'),text="PAID/UNPAID",bd=16,anchor='w',fg="steel blue",bg="powder blue")
                pn.grid(row=11,column=0)
                Radiobutton(f1,font=('arial',16,'bold'),text="PAID",fg="blue",variable = pd_ud,value = "PAID",bd=10,justify = 'right',).place(x=200,y=500)
                Radiobutton(f1,font=('arial',16,'bold'),text="UNPAID",fg = "blue",variable = pd_ud,value = "UNPAID",bd=10).place(x=312,y=500)
                
                a_new = Label(f1,font=('arial',18,'bold'),text="AMOUNT",bd=16,anchor='w',fg="steel blue",bg="powder blue")
                a_new.grid(row=12,column=0)
                amount_new = Entry(f1,font=('arial',16,'bold'), bd=10,insertwidth=4,fg="steel blue",bg="powder blue",justify = 'right')
                amount_new.grid(row=12,column=1)
                amount_new.focus_set()
                
                def insertnew():
                    sr_new.focus_set()
                    def clearnew():
    
                    # clear the content of text entry box
    
                        sr_new.delete(0, END)
                        
                        amount_new.delete(0, END)
                    pdud=pd_ud.get()
                    amount_new1 = amount_new.get()
                    
                    
                    for i in range(2,sheet.max_row+1,1):
                        cell_sheet=sheet.cell(row=i,column=1)
                        
                        if cell_sheet.value == sr_new1:
                            if pdud != "":
                                sheet.cell(row=i,column=4).value = pd_ud.get()
                            if amount_new1 != "":
                                sheet.cell(row=i,column=3).value = amount_new1
                            
                            wb.save('/home/vickyvirus/VICKYVIRUS PROGRAMS/shivdatt2018.xlsx')
                            clearnew()
                            
                            
                            

                            
                            
                             
        
                            
                            
                         
                insertdata = Button(f1,font=('arial',14,'bold'),text="INSERT",fg="black",bg="red",bd=16,width=4,command = insertnew ).grid(row=13,column=1,sticky="nsew")
                 
                    
                
                
             
                
                
                
                        
    search = Button(f1,font=('arial',14,'bold'),text="SEARCH",fg="black",bg="red",bd=16,width=4,command = edit).grid(row=9,column=1,sticky="nsew")
      
      
    
paid_unpaid = StringVar()




s = Label(f1,font=('arial',18,'bold'),text="SR",bd=16,anchor='w',fg="steel blue",bg="powder blue")
s.grid(row=0,column=0)
sr = Label(f1,font=('arial',20,'bold'),bd=10,bg="powder blue",fg="blue",justify = 'right')
sr.grid(row=0,column=1)

n = Label(f1,font=('arial',18,'bold'),text="NAME",bd=16,anchor='w',fg="steel blue",bg="powder blue")
n.grid(row=1,column=0)
name = Entry(f1,font=('arial',16,'bold'),bd=10,insertwidth=4,fg="blue",bg="powder blue",justify = 'right')
name.grid(row=1,column=1)

a = Label(f1,font=('arial',18,'bold'),text="AMOUNT",bd=16,anchor='w',fg="steel blue",bg="powder blue")
a.grid(row=2,column=0)
amount = Entry(f1,font=('arial',16,'bold'), bd=10,insertwidth=4,fg="steel blue",bg="powder blue",justify = 'right')
amount.grid(row=2,column=1)

p = Label(f1,font=('arial',18,'bold'),text="PAID/UNPAID",bd=16,anchor='w',fg="steel blue",bg="powder blue")
p.grid(row=3,column=0)
Radiobutton(f1,font=('arial',16,'bold'),text="PAID",fg="blue",variable = paid_unpaid,value = "PAID",bd=10,justify = 'right',).place(x=208,y=191)
Radiobutton(f1,font=('arial',16,'bold'),text="UNPAID",fg = "blue",variable = paid_unpaid,value = "UNPAID",bd=10).place(x=320,y=191)

excel()
submit = Button(f1,font=('arial',14,'bold'),text="Submit",fg="black",bg = "red",bd=10,width=2,command= insert,justify='right').grid(row=4,column=1,sticky="nsew")

total = Button(f1,font=('arial',14,'bold'),text="TOTAL",fg="black",bg="red",bd=16,width=4,command = total).grid(row=4,column=2,sticky="nsew")
totalamount = Label(f1,font=('arial',20,'bold'), bd=10,fg="green",bg="powder blue")  
totalamount.grid(row=0,column=3)
paidamount = Label(f1,font=('arial',20,'bold'), bd=10,fg="green",bg="powder blue")  
paidamount.grid(row=1,column=3)
unpaidamount = Label(f1,font=('arial',20,'bold'), bd=10,fg="green",bg="powder blue")  
unpaidamount.grid(row=2,column=3)


#=======================change======
change = Button(f1,font=('arial',14,'bold'),text="CHANGE",fg="black",bg="red",bd=16,width=4,command = change).grid(row=4,column=3,sticky="nsew")
exit_software = Button(f1,font=('arial',14,'bold'),text="QUIT",fg="black",bg="red",bd=16,width=4,command = exit).grid(row=4,column=4,sticky="nsew")
root.mainloop()


#=====================extrafeatures===========================================






        


